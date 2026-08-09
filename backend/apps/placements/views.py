import csv
import io
import re
import threading
import time
from datetime import timedelta
from math import ceil

from django.contrib.auth import get_user_model
from django.db.models import Count, Q, Sum
from django.db.models.functions import TruncDate
from django.utils import timezone
from rest_framework import permissions, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet

from apps.core.models import Notification, SiteSetting
from apps.core.permissions import IsSuperAdmin
from apps.core.throttles import AiRateThrottle
from apps.core.utils import log_audit, notify

from .ai import AiError, REASONING_BUDGETS, ai_json, ai_plain_text
from .models import AiUsageLog, Drive
from .resume_ai import maybe_refresh_drive_matches
from .serializers import DriveSerializer

User = get_user_model()

# Admins and CRs may post/manage drives; faculty only if the admin gave them
# the Placement portal access (some faculty are resume-review only).
_WRITE_ROLES = {User.Role.SUPER_ADMIN, User.Role.CR}

# Monthly AI credit budget (in tokens) the super admin sets on the AI Usage page.
_AI_BUDGET_KEY = "ai_monthly_budget_tokens"

# Reasonable input caps so one pasted block can't burn the whole AI quota.
_MAX_PASTE_CHARS = 10_000
_MAX_QUESTION_CHARS = 1_000

# Roll-number-ish tokens: optional leading digits (year) + 2-6 letters + digits
# (21CSE01, 21CSE1A05, 211FA04001…). The leading year digits are optional.
_ROLL_TOKEN_RE = re.compile(r"\b\d{0,4}[A-Za-z]{2,6}\d{1,6}[A-Za-z0-9]{0,6}\b")
_ROLL_HEADER_RE = re.compile(r"roll|reg(?:istration)?\s*no|register|hall\s*ticket|enroll|college\s*id", re.I)
_ELIG_HEADER_RE = re.compile(r"eligib|criteria|branch(?:es)?\s*allowed|passout|cgpa", re.I)

_EXTRACT_PROMPT = """\
You extract placement drive details from forwarded WhatsApp/college messages into a JSON object.
Return ONLY valid JSON (no markdown, no code fences, no comments) with exactly these keys - use
null for missing values and empty arrays for lists that are not mentioned:
- company_name (string)
- job_role (string)
- job_type (string, e.g. "Full-time", "Internship", "Contract")
- location (string)
- work_mode (string, e.g. "On-site", "Remote", "Hybrid")
- package (string, e.g. "6 LPA" or "12k/month" or "6-8 LPA")
- minimum_package (string, the lower bound when the message gives a range like "6-8 LPA")
- maximum_package (string, the upper bound when the message gives a range)
- eligibility (string, free-text eligibility criteria)
- eligible_branches (array of branch names, e.g. ["CSE", "IT"])
- minimum_cgpa (string, e.g. "6.5" - null if not mentioned)
- maximum_backlogs (string, e.g. "0" - null if not mentioned)
- passing_year (array of pass-out years, e.g. ["2025", "2026"])
- eligible_roll_numbers (string, comma-separated roll numbers if the message lists them)
- selection_process (string, e.g. "Aptitude -> Technical -> HR")
- application_start_date (string in YYYY-MM-DD if mentioned, else null)
- application_deadline (string in YYYY-MM-DD; convert phrases like "last date: 15th Aug" or
  "apply before 12/08/2026"; null if not mentioned)
- apply_link (string, the apply/registration URL if mentioned)
- job_description (string, a short 1-2 sentence summary of the drive)
- required_skills (array of skills, e.g. ["Python", "SQL"])
- important_instructions (string, any special notes students must know)
- company_description (string, what the company does - only if the message says)
Never invent information that is not in the message.
"""

# RAG-grounded chat prompt: the drive details arrive as grounding documents
# (via the RAG service or prompt injection), so the model answers using ONLY
# the real drive data - never invented facts.
_RAG_CHAT_PROMPT = """\
You are the placement assistant for a college. Students ask about placement drives.
Answer using ONLY the drive details in the documents. Answer in clear, friendly, concise
English (2-4 sentences). If a drive is not listed or a detail is missing, say so honestly and
suggest contacting the placement cell. For eligibility questions, compare the student's branch,
section, pass-out year and roll number with each drive's eligibility and eligible roll numbers,
and say whether they qualify. If the question needs current outside information (news, the
company's website, new openings) that is not in the documents, say clearly that you don't
have live web access and suggest contacting the placement cell - never invent it.
"""


def _normalize_extraction(extracted: dict) -> dict:
    """Map the structured Nemotron fields onto the drive form's field names.

    The extraction returns the rich 22-field structure; this builds the legacy
    form keys (company_name, role, package, drive_link, last_date_to_apply,
    description, eligibility, eligible_roll_numbers) so the existing frontend
    auto-fill keeps working, folding the extra structured details into the
    description/eligibility text so nothing is lost.
    """

    def pick(*keys) -> str:
        for key in keys:
            value = extracted.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()
            if isinstance(value, list) and value:
                return ", ".join(str(item) for item in value)
        return ""

    package = pick("package")
    if not package and pick("minimum_package") and pick("maximum_package"):
        package = f"{pick('minimum_package')}-{pick('maximum_package')}"

    normalized = {
        "company_name": pick("company_name"),
        "role": pick("job_role", "role"),
        "location": pick("location"),
        "package": package,
        "drive_link": pick("apply_link", "drive_link"),
        "last_date_to_apply": pick("application_deadline", "last_date_to_apply"),
    }

    detail_parts = [pick("job_description")]
    extra_parts = []
    if pick("job_type"):
        extra_parts.append(f"Job type: {pick('job_type')}")
    if pick("work_mode"):
        extra_parts.append(f"Work mode: {pick('work_mode')}")
    if pick("selection_process"):
        extra_parts.append(f"Selection process: {pick('selection_process')}")
    if pick("required_skills"):
        extra_parts.append(f"Skills required: {pick('required_skills')}")
    if pick("important_instructions"):
        extra_parts.append(pick("important_instructions"))
    if pick("company_description"):
        extra_parts.append(pick("company_description"))
    if extra_parts:
        detail_parts.append(" ".join(extra_parts))
    normalized["description"] = " | ".join(p for p in detail_parts if p)

    elig_parts = [pick("eligibility")]
    if pick("eligible_branches"):
        elig_parts.append(f"Eligible branches: {pick('eligible_branches')}")
    if pick("minimum_cgpa"):
        elig_parts.append(f"Minimum CGPA: {pick('minimum_cgpa')}")
    if pick("maximum_backlogs"):
        elig_parts.append(f"Maximum backlogs: {pick('maximum_backlogs')}")
    if pick("passing_year"):
        elig_parts.append(f"Passing year: {pick('passing_year')}")
    normalized["eligibility"] = "; ".join(p for p in elig_parts if p)
    normalized["eligible_roll_numbers"] = pick("eligible_roll_numbers")
    return normalized


# ---------------------------------------------------------------------------
# Fast DB answers & caching for the AI chat (no LLM for simple questions)
# ---------------------------------------------------------------------------
_QA_CACHE_TTL_SECONDS = 300  # 5 minutes
_qa_cache: dict = {}
_qa_lock = threading.Lock()

# Questions whose answer depends on the asking student are never cached.
_STUDENT_SPECIFIC_RE = re.compile(
    r"eligib|can i|can a|am i|for me|qualif|my roll|apply to|i have", re.I
)

_QUICK_PACKAGE_RE = re.compile(r"package|salary|ctc|lpa|stipend|\bpay\b", re.I)
_QUICK_DEADLINE_RE = re.compile(
    r"last date|deadline|apply by|apply before|clos(?:e|ing)|due date|when.*apply", re.I
)
_QUICK_LINK_RE = re.compile(
    r"apply link|application link|how.*apply|where.*apply|register", re.I
)
_QUICK_ELIGIBLE_RE = re.compile(r"eligib|can a|can i|am i", re.I)


def _normalize_question(question: str) -> str:
    return re.sub(r"\s+", " ", question.strip().lower())


def _cached_answer(cache_key: tuple, question: str, provider):
    """Serve repeated identical questions from a short in-memory cache.

    Student-specific questions (eligibility, "can I apply") always go to the
    provider - caching a personal answer would be wrong. Everything else is
    cached for 5 minutes so ~1000 students asking the same thing don't all
    burn an LLM call.
    """
    if _STUDENT_SPECIFIC_RE.search(question):
        return provider()
    now = time.monotonic()
    with _qa_lock:
        hit = _qa_cache.get(cache_key)
        if hit and now - hit[0] < _QA_CACHE_TTL_SECONDS:
            return hit[1]
    answer = provider()
    with _qa_lock:
        _qa_cache[cache_key] = (time.monotonic(), answer)
        if len(_qa_cache) > 2000:
            for key in [
                k for k, (ts, _) in _qa_cache.items()
                if now - ts > _QA_CACHE_TTL_SECONDS
            ]:
                _qa_cache.pop(key, None)
    return answer


def _quick_drive_answer(drive, user, question):
    """Answer simple fact questions straight from the drive row - no LLM."""
    q = question.lower()
    if _QUICK_DEADLINE_RE.search(q):
        if not drive.last_date_to_apply:
            return (
                f"No apply-by date was shared for {drive.company_name} - "
                "contact the placement cell for details."
            )
        return (
            f"The last date to apply for {drive.company_name} is "
            f"{drive.last_date_to_apply}."
        )
    if _QUICK_LINK_RE.search(q):
        if drive.drive_link:
            return (
                f"You can apply for {drive.company_name} here: {drive.drive_link}"
            )
        return (
            f"No apply link was shared for {drive.company_name} yet - "
            "contact the placement cell."
        )
    if _QUICK_PACKAGE_RE.search(q):
        return (
            f"{drive.company_name} offers {drive.package}."
            if drive.package
            else f"The package for {drive.company_name} wasn't mentioned."
        )
    if _QUICK_ELIGIBLE_RE.search(q):
        if (
            user.is_student
            and user.roll_number
            and user.roll_number.strip().upper() in drive.eligible_rolls()
        ):
            return (
                f"Yes! Your roll number ({user.roll_number}) is in the "
                f"pre-approved eligible list for {drive.company_name}. Good luck!"
            )
        text = (
            f"Eligibility for {drive.company_name}: "
            f"{drive.eligibility or 'not mentioned yet'}."
        )
        if drive.eligible_roll_numbers:
            text += " The company also sent a pre-approved roll-number list."
        return text
    return None


def _mentioned_drive(drives, question):
    """Return the first open drive whose company name appears in the question."""
    q = question.lower()
    for d in drives:
        if d.company_name and d.company_name.lower() in q:
            return d
    return None


def _student_line(user) -> str:
    """Compact student profile used as AI chat context."""
    branch = user.branch.name if user.branch_id else "branch not set"
    section = f"/ Sec {user.section.name}" if user.section_id else ""
    batch = f", passout {user.passout_year}" if getattr(user, "passout_year", None) else ""
    return f"{user.full_name} (roll {user.roll_number}, {branch}{section}{batch})"


def _drive_recipients():
    """All active students - they are the ones applying to drives."""
    return User.objects.filter(is_active=True, role=User.Role.STUDENT)


def _log_ai_usage(user, action):
    """Build a usage callback that records one AI call's token counts."""

    def callback(prompt_tokens: int, completion_tokens: int) -> None:
        try:
            AiUsageLog.objects.create(
                user=user, action=action,
                prompt_tokens=prompt_tokens, completion_tokens=completion_tokens,
            )
        except Exception:  # pragma: no cover - usage tracking must never break the request
            pass

    return callback


def _ai_budget_tokens() -> int:
    try:
        setting = SiteSetting.objects.filter(key=_AI_BUDGET_KEY).first()
        if setting:
            return max(0, int(str(setting.value) or 0))
    except (TypeError, ValueError):
        pass
    return 0


def _drive_preview(instance) -> str:
    line = instance.role or instance.location or ""
    preview = instance.company_name
    if line:
        preview += f" · {line}"
    if instance.package:
        preview += f" · {instance.package}"
    # Keep well under the Notification.message (max_length=500) cap.
    return preview[:300]


def _parse_eligibility_file(uploaded) -> dict:
    """Extract roll numbers (and an optional eligibility column) from .xlsx/.csv.

    Finds the roll-number column by header name; falls back to scanning every
    cell for roll-number-shaped tokens. Also picks up an eligibility-style
    column (eligibility/branch/cgpa/passout) if one exists.
    """
    name = (uploaded.name or "").lower()
    raw = uploaded.read()
    if name.endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        rows = [
            [str(c).strip() if c is not None else "" for c in row]
            for sheet in wb.worksheets
            for row in sheet.iter_rows(values_only=True)
        ]
    elif name.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
        rows = [[c.strip() for c in row] for row in rows]
    else:
        raise ValueError("Please upload an .xlsx or .csv file.")

    rows = [row for row in rows if any(row)]

    # 1) Try to find the roll-number column by its header.
    roll_col = None
    elig_col = None
    for row in rows[:10]:
        for i, cell in enumerate(row):
            if roll_col is None and _ROLL_HEADER_RE.search(cell):
                roll_col = i
            if elig_col is None and _ELIG_HEADER_RE.search(cell):
                elig_col = i
        if roll_col is not None and elig_col is not None:
            break

    rolls: set[str] = set()
    elig_values: list[str] = []
    for row in rows[1:] if (roll_col is not None or elig_col is not None) else rows:
        if roll_col is not None and roll_col < len(row):
            token = row[roll_col].strip()
            if token and _ROLL_TOKEN_RE.fullmatch(token):
                rolls.add(token.upper())
        elif roll_col is None:
            # No header found - regex every cell.
            for cell in row:
                for match in _ROLL_TOKEN_RE.finditer(cell):
                    rolls.add(match.group(0).upper())
        if elig_col is not None and elig_col < len(row):
            value = row[elig_col].strip()
            if value and value not in elig_values:
                elig_values.append(value)

    return {
        "roll_numbers": ", ".join(sorted(rolls)),
        "count": len(rolls),
        "eligibility": "; ".join(elig_values[:20])[:1000],
    }


class _CanWriteDrives(permissions.BasePermission):
    """Authenticated users may read drives; only admins, CRs and faculty with
    the Placement portal access may write."""

    def has_permission(self, request, view):
        user = request.user
        if not (user and user.is_authenticated):
            return False
        if request.method in permissions.SAFE_METHODS:
            return True
        if user.is_faculty:
            return user.has_placement_portal
        return user.role in _WRITE_ROLES


class DriveViewSet(ModelViewSet):
    """Placement drives. Open drives expire at the last date to apply and are
    hard-deleted 30 days after expiry."""

    serializer_class = DriveSerializer
    permission_classes = [_CanWriteDrives]
    # Drives are a small, curated list - no pagination wrapper for the tabs.
    pagination_class = None
    MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB cap for eligibility sheets

    def get_serializer_context(self):
        """Pre-load the student's AI match snapshot once per request so the
        serializer never queries per drive."""
        context = super().get_serializer_context()
        user = self.request.user
        resume_match = {}
        if user and user.is_authenticated and user.is_student:
            resume = getattr(user, "resume", None)
            if resume and isinstance(resume.ai_match, dict):
                resume_match = resume.ai_match
        context["resume_match"] = resume_match
        return context

    def get_throttles(self):
        # The AI calls cost real API quota - throttle them per user.
        if self.action in ("ai_extract", "ai_chat", "ai_ask"):
            return [AiRateThrottle()]
        return super().get_throttles()

    def get_permissions(self):
        # The AI chat answers eligibility questions for students too.
        if self.action in ("ai_chat", "ai_ask", "my_ai_usage"):
            return [IsAuthenticated()]
        # Credit usage is admin-only.
        if self.action in ("ai_usage", "ai_budget"):
            return [IsSuperAdmin()]
        return [_CanWriteDrives()]

    def get_queryset(self):
        # Lazy cleanup: drop anything past its 30-day grace period. Drives
        # without a last date (NULL) never expire and are never cleaned up.
        cutoff = timezone.localdate() - timedelta(days=30)
        Drive.objects.filter(last_date_to_apply__lt=cutoff).delete()

        qs = Drive.objects.select_related("posted_by").all()
        # Only the list tab filters by status - detail/update/delete must see
        # every drive regardless of its current status.
        if self.action == "list":
            drive_status = (self.request.query_params.get("status") or "open").lower()
            today = timezone.localdate()
            if drive_status == "open":
                # Drives with no deadline stay open forever.
                qs = qs.filter(Q(last_date_to_apply__isnull=True) | Q(last_date_to_apply__gte=today))
            elif drive_status == "expired":
                qs = qs.filter(last_date_to_apply__lt=today)
        return qs

    def check_object_permissions(self, request, obj):
        super().check_object_permissions(request, obj)
        user = request.user
        if request.method in ("PUT", "PATCH", "DELETE"):
            # Only the poster or a super admin may edit/delete a drive.
            if not (user.is_super_admin or obj.posted_by_id == user.id):
                self.permission_denied(
                    request,
                    message="You can only edit or delete drives you posted.",
                )

    # ------------------------------------------------------------ AI actions

    @action(detail=False, methods=["post"])
    def ai_extract(self, request):
        """Turn a pasted WhatsApp/college message into structured drive fields."""
        text = (request.data.get("text") or "").strip()
        if len(text) < 10:
            return Response(
                {"detail": "Paste at least a few lines of the drive message first."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if len(text) > _MAX_PASTE_CHARS:
            return Response(
                {"detail": f"That paste is too long (max {_MAX_PASTE_CHARS} characters)."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            extracted = ai_json(
                _EXTRACT_PROMPT, text, max_tokens=2048,
                reasoning_budget=REASONING_BUDGETS["extract"],
                usage_callback=_log_ai_usage(request.user, AiUsageLog.Action.EXTRACT),
            )
        except AiError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_502_BAD_GATEWAY)
        if not extracted:
            return Response(
                {"detail": "The AI could not understand this message. Try a longer paste."},
                status=status.HTTP_422_UNPROCESSABLE_ENTITY,
            )
        # The rich structured fields are returned alongside the legacy form
        # keys so the frontend auto-fill keeps working unchanged.
        payload = {**extracted, **_normalize_extraction(extracted)}
        return Response(payload)

    @action(detail=False, methods=["post"])
    def parse_eligibility(self, request):
        """Upload an .xlsx/.csv and get the eligible roll numbers (and an
        eligibility column if present) back as text."""
        uploaded = request.FILES.get("file")
        if not uploaded:
            return Response({"detail": "Attach an .xlsx or .csv file."}, status=400)
        if uploaded.size > self.MAX_UPLOAD_BYTES:
            return Response(
                {"detail": "That file is too large (max 5 MB)."}, status=400
            )
        try:
            result = _parse_eligibility_file(uploaded)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=400)
        if result["count"] == 0:
            return Response(
                {"detail": "No roll numbers found in that file.", "roll_numbers": "", "count": 0, "eligibility": ""},
                status=200,
            )
        return Response(result)

    @action(detail=True, methods=["post"])
    def ai_ask(self, request, pk=None):
        """Answer a question about one specific drive (the 'Ask AI' button).

        Simple factual questions (package, last date, apply link, eligibility)
        are answered straight from the drive row - no LLM call, so they stay
        instant and free. Complex questions go to Nemotron with the drive +
        student profile as context, and identical questions are served from a
        short-lived cache.
        """
        question = (request.data.get("question") or "").strip()
        if len(question) < 3:
            return Response({"detail": "Ask a question about this drive."}, status=400)
        if len(question) > _MAX_QUESTION_CHARS:
            return Response({"detail": "That question is too long."}, status=400)

        drive = self.get_object()
        user = request.user

        quick = _quick_drive_answer(drive, user, question)
        if quick is not None:
            return Response({"answer": quick})

        # The drive's real details are the RAG grounding documents - the model
        # may only answer from these (eligibility questions also see the
        # student's own profile in the system prompt).
        documents = [
            f"Drive: {drive.company_name} ({drive.role or 'role not mentioned'}, "
            f"{drive.package or 'package not mentioned'}, {drive.location or 'location not mentioned'}).\n"
            f"Details: {(drive.description or 'not mentioned')[:300]}\n"
            f"Eligibility: {(drive.eligibility or 'not mentioned')[:300]}\n"
            f"Eligible rolls: {drive.eligible_roll_numbers or 'not listed'}\n"
            f"Last date to apply: {drive.last_date_to_apply or 'not announced'}\n"
            f"Apply link: {drive.drive_link or 'not provided'}"
        ]
        system = _RAG_CHAT_PROMPT + f"\nAsking student: {_student_line(user)}."

        def ask():
            return ai_plain_text(
                system, question, max_tokens=800,
                reasoning_budget=REASONING_BUDGETS["chat"],
                usage_callback=_log_ai_usage(request.user, AiUsageLog.Action.ASK),
                documents=documents,
            )

        try:
            answer = _cached_answer(
                ("ai_ask", drive.id, _normalize_question(question)), question, ask
            )
        except AiError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_502_BAD_GATEWAY)
        return Response({"answer": answer})

    @action(detail=False, methods=["post"])
    def ai_chat(self, request):
        """Answer a student's placement/eligibility question with AI context.

        When the question names a drive and asks a simple fact (deadline,
        package, apply link, eligibility) it is answered straight from the
        database - no LLM call. Complex questions go to Nemotron with the open
        drives + student profile as context, and repeated identical questions
        are served from a short-lived cache.
        """
        question = (request.data.get("question") or "").strip()
        if len(question) < 3:
            return Response(
                {"detail": "Ask a question about the drives."}, status=400
            )
        if len(question) > _MAX_QUESTION_CHARS:
            return Response({"detail": "That question is too long."}, status=400)

        drives = list(
            Drive.objects.filter(last_date_to_apply__gte=timezone.localdate())
            .select_related("posted_by")
            .order_by("-created_at")[:12]
        )
        user = request.user

        # Quick DB answer when the question names a drive and asks a fact.
        mentioned = _mentioned_drive(drives, question)
        if mentioned is not None:
            quick = _quick_drive_answer(mentioned, user, question)
            if quick is not None:
                return Response({"answer": quick})

        # Every open drive becomes a grounding document for RAG - the model may
        # only answer from these (plus the student's own profile).
        documents = []
        for d in drives:
            documents.append(
                f"Drive: {d.company_name} ({d.role or 'role not mentioned'}, "
                f"{d.package or 'package not mentioned'}, {d.location or 'location not mentioned'}).\n"
                f"Details: {(d.description or 'not mentioned')[:300]}\n"
                f"Eligibility: {(d.eligibility or 'not mentioned')[:300]}\n"
                f"Eligible rolls: {d.eligible_roll_numbers or 'not listed'}\n"
                f"Last date to apply: {d.last_date_to_apply or 'not announced'}\n"
                f"Apply link: {d.drive_link or 'not provided'}"
            )
        if not documents:
            documents.append("There are no open drives right now.")
        system = _RAG_CHAT_PROMPT + f"\nAsking student: {_student_line(user)}."

        def chat():
            return ai_plain_text(
                system, question, max_tokens=800,
                reasoning_budget=REASONING_BUDGETS["chat"],
                usage_callback=_log_ai_usage(request.user, AiUsageLog.Action.CHAT),
                documents=documents,
            )

        try:
            answer = _cached_answer(
                ("ai_chat", _normalize_question(question)), question, chat
            )
        except AiError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_502_BAD_GATEWAY)
        return Response({"answer": answer})

    # --------------------------------------------------- AI credits (admin)

    @action(detail=False, methods=["get"])
    def ai_usage(self, request):
        """Totals + per-user AI credit usage, with the monthly budget."""
        # Retention: drop logs older than 90 days so the table never grows
        # unbounded (same lazy-cleanup pattern as expired drives).
        AiUsageLog.objects.filter(
            created_at__lt=timezone.now() - timedelta(days=90)
        ).delete()
        logs = AiUsageLog.objects.all()
        total_calls = logs.count()
        totals = logs.aggregate(
            prompt_tokens=Sum("prompt_tokens"), completion_tokens=Sum("completion_tokens"),
        )
        used_tokens = int(totals["prompt_tokens"] or 0) + int(totals["completion_tokens"] or 0)

        rows = (
            logs.values("user_id")
            .annotate(
                calls=Count("id"),
                prompt_tokens=Sum("prompt_tokens"),
                completion_tokens=Sum("completion_tokens"),
            )
            .order_by("-calls")
        )
        users = {
            u.id: u
            for u in User.objects.filter(id__in=[r["user_id"] for r in rows])
        }
        per_user = []
        for row in rows:
            user = users.get(row["user_id"])
            user_tokens = int(row["prompt_tokens"] or 0) + int(row["completion_tokens"] or 0)
            per_user.append({
                "user_id": row["user_id"],
                "name": user.full_name if user else "Deleted user",
                "roll_number": user.roll_number if user else "—",
                "role": user.role if user else "",
                "calls": row["calls"],
                "prompt_tokens": int(row["prompt_tokens"] or 0),
                "completion_tokens": int(row["completion_tokens"] or 0),
                "total_tokens": user_tokens,
            })

        # Daily breakdown (last 30 days) so the admin sees usage day by day.
        daily_rows = {
            row["day"]: row
            for row in logs.filter(created_at__gte=timezone.now() - timedelta(days=30))
            .annotate(day=TruncDate("created_at"))
            .values("day")
            .annotate(
                calls=Count("id"),
                tokens=Sum("prompt_tokens") + Sum("completion_tokens"),
            )
            .order_by("day")
        }
        today = timezone.localdate()
        daily = []
        for i in range(29, -1, -1):
            day = today - timedelta(days=i)
            row = daily_rows.get(day)
            daily.append({
                "date": day.isoformat(),
                "calls": row["calls"] if row else 0,
                "tokens": int(row["tokens"] or 0) if row else 0,
            })

        budget = _ai_budget_tokens()
        return Response({
            "totals": {
                "calls": total_calls,
                "prompt_tokens": int(totals["prompt_tokens"] or 0),
                "completion_tokens": int(totals["completion_tokens"] or 0),
                "used_tokens": used_tokens,
            },
            "daily": daily,
            "per_user": per_user,
            "budget_tokens": budget,
            "remaining_tokens": max(0, budget - used_tokens) if budget else None,
            "percent_used": round(used_tokens / budget * 100, 1) if budget else None,
        })

    @action(detail=False, methods=["get"])
    def my_ai_usage(self, request):
        """Your own AI credit usage - visible to every authenticated user."""
        logs = AiUsageLog.objects.filter(user=request.user)
        totals = logs.aggregate(
            prompt_tokens=Sum("prompt_tokens"), completion_tokens=Sum("completion_tokens"),
        )
        used_tokens = int(totals["prompt_tokens"] or 0) + int(totals["completion_tokens"] or 0)
        recent = [
            {
                "action": log.action,
                "action_label": log.get_action_display(),
                "prompt_tokens": log.prompt_tokens,
                "completion_tokens": log.completion_tokens,
                "total_tokens": log.total_tokens,
                "created_at": log.created_at.isoformat(),
            }
            for log in logs[:20]
        ]
        return Response({
            "calls": logs.count(),
            "used_tokens": used_tokens,
            "credits": ceil(used_tokens / 1000),
            "recent": recent,
        })

    @action(detail=False, methods=["post"])
    def ai_budget(self, request):
        """Set the monthly AI credit budget (in tokens) for the usage page."""
        try:
            budget = int(request.data.get("budget_tokens") or 0)
        except (TypeError, ValueError):
            return Response({"detail": "Enter a valid number of tokens."}, status=400)
        if budget < 0:
            return Response({"detail": "The budget can't be negative."}, status=400)
        SiteSetting.objects.update_or_create(
            key=_AI_BUDGET_KEY, defaults={"value": str(budget)}
        )
        log_audit(request.user, "UPDATE", "SiteSetting", _AI_BUDGET_KEY,
                  {"ai_monthly_budget_tokens": budget}, request)
        return Response({"budget_tokens": budget})

    # ------------------------------------------------------------ lifecycle

    def perform_create(self, serializer):
        instance = serializer.save(posted_by=self.request.user)
        log_audit(
            self.request.user, "CREATE", "Drive", instance.id,
            {"company_name": instance.company_name, "role": instance.role},
            self.request,
        )
        notify(
            _drive_recipients(),
            Notification.Kind.DRIVE,
            f"New drive: {instance.company_name}",
            _drive_preview(instance),
            f"/placements/{instance.id}",
        )
        # Refresh the AI match for this new drive across already-analyzed
        # resumes (background, best-effort) so students see their match score
        # without re-running anything - never blocks the drive post.
        try:
            maybe_refresh_drive_matches(instance, self.request.user)
        except Exception:  # pragma: no cover - the post must always succeed
            pass

    def perform_update(self, serializer):
        instance = serializer.save()
        log_audit(
            self.request.user, "UPDATE", "Drive", instance.id,
            {"company_name": instance.company_name}, self.request,
        )

    def perform_destroy(self, instance):
        log_audit(
            self.request.user, "DELETE", "Drive", instance.id,
            {"company_name": instance.company_name}, self.request,
        )
        instance.delete()
